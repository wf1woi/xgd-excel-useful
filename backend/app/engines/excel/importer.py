import logging
from collections.abc import Callable
from io import BytesIO

from openpyxl import load_workbook

from app.engines.excel.structure import build_detected_columns, column_letter_to_index
from app.engines.excel.worksheet_bounds import resolve_worksheet_bounds
from app.services.fixed_field import resolve_fixed_field_values


logger = logging.getLogger(__name__)


class ExcelImportEngine:
    def extract_rows(
        self,
        content: bytes,
        sheet_name: str,
        header_row_index: int,
        data_start_row_index: int,
        data_end_column: str,
        ignore_empty_row: bool,
        fixed_fields: list[dict[str, str | bool | None]] | None = None,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> tuple[list[dict[str, str | None]], list[dict[str, str | int | None]]]:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise ValueError("指定的 sheet 不存在")

        worksheet = workbook[sheet_name]
        sheet_max_rows, _sheet_max_columns = resolve_worksheet_bounds(worksheet, logger)
        resolved_fixed_fields = resolve_fixed_field_values(worksheet, fixed_fields)
        total_candidate_rows = max(sheet_max_rows - data_start_row_index + 1, 0)
        end_column_index = column_letter_to_index(data_end_column) + 1
        header_values = next(
            worksheet.iter_rows(
                min_row=header_row_index,
                max_row=header_row_index,
                max_col=end_column_index,
                values_only=True,
            ),
            tuple(),
        )
        sample_values = next(
            worksheet.iter_rows(
                min_row=data_start_row_index,
                max_row=data_start_row_index,
                max_col=end_column_index,
                values_only=True,
            ),
            tuple(),
        )
        detected_columns = build_detected_columns(
            rows=[list(header_values), list(sample_values)],
            header_row_index=1,
            data_start_row_index=2,
            data_end_column=data_end_column,
        )

        rows: list[dict[str, str | None]] = []
        processed_rows = 0
        for row in worksheet.iter_rows(
            min_row=data_start_row_index,
            max_col=end_column_index,
            values_only=True,
        ):
            processed_rows += 1
            values = list(row)
            if ignore_empty_row and all(value in (None, "") for value in values):
                if progress_callback and (processed_rows == 1 or processed_rows % 1000 == 0):
                    progress_callback(processed_rows, total_candidate_rows)
                continue

            record: dict[str, str | None] = {}
            for column in detected_columns:
                cell_value = values[column.column_index] if column.column_index < len(values) else None
                record[column.field_name] = None if cell_value in (None, "") else str(cell_value).strip()
            for fixed_field in resolved_fixed_fields:
                record[fixed_field["field_key"]] = fixed_field["field_value"] or None
            rows.append(record)
            if progress_callback and (processed_rows == 1 or processed_rows % 1000 == 0):
                progress_callback(processed_rows, total_candidate_rows)

        if progress_callback:
            progress_callback(processed_rows, total_candidate_rows)

        return rows, [
            {
                "column_index": column.column_index,
                "column_letter": column.column_letter,
                "header_name": column.header_name,
                "field_name": column.field_name,
                "sample_value": column.sample_value,
            }
            for column in detected_columns
        ]
