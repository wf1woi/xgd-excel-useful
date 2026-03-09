import logging
from io import BytesIO

from openpyxl import load_workbook

from app.engines.excel.structure import build_detected_columns
from app.engines.excel.worksheet_bounds import resolve_worksheet_bounds


logger = logging.getLogger(__name__)


class ExcelPreviewEngine:
    def preview(
        self,
        content: bytes,
        sheet_name: str | None = None,
        max_rows: int = 50,
        max_columns: int = 26,
    ) -> dict:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )
        sheet_names = workbook.sheetnames
        selected_sheet_name = sheet_name or workbook.active.title
        if selected_sheet_name not in sheet_names:
            raise ValueError("指定的 sheet 不存在")

        worksheet = workbook[selected_sheet_name]
        sheet_max_rows, sheet_max_columns = resolve_worksheet_bounds(worksheet, logger)
        rows: list[list[str | int | float | bool | None]] = []

        for row_index, row in enumerate(
            worksheet.iter_rows(max_row=max_rows, max_col=max_columns, values_only=True),
            start=1,
        ):
            values = list(row)
            if row_index > max_rows:
                break
            rows.append(values)

        detected_columns = build_detected_columns(
            rows=rows,
            header_row_index=1,
            data_start_row_index=2,
            data_end_column=self._infer_end_column(rows),
        )

        return {
            "sheet_names": sheet_names,
            "selected_sheet_name": selected_sheet_name,
            "max_rows": max_rows,
            "max_columns": max_columns,
            "sheet_max_rows": sheet_max_rows,
            "sheet_max_columns": sheet_max_columns,
            "is_truncated_rows": sheet_max_rows > max_rows,
            "is_truncated_columns": sheet_max_columns > max_columns,
            "detected_columns": [
                {
                    "column_index": column.column_index,
                    "column_letter": column.column_letter,
                    "header_name": column.header_name,
                    "field_name": column.field_name,
                    "sample_value": column.sample_value,
                }
                for column in detected_columns
            ],
            "rows": rows,
        }

    def _infer_end_column(self, rows: list[list[str | int | float | bool | None]]) -> str:
        max_column_index = 0
        for row in rows:
            for column_index in range(len(row) - 1, -1, -1):
                if row[column_index] not in (None, ""):
                    max_column_index = max(max_column_index, column_index)
                    break

        from app.engines.excel.structure import to_column_letter

        return to_column_letter(max_column_index)
