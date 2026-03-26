from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.export_preview import WorkbookPreview


logger = logging.getLogger(__name__)


class ExportExcelService:
    def __init__(self, export_dir: Path) -> None:
        self.export_dir = export_dir

    def build_file(self, workbook_preview: WorkbookPreview) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)

        header_fill = PatternFill(fill_type="solid", fgColor="1F4D43")
        header_font = Font(color="FFFFFF", bold=True)
        summary_fill = PatternFill(fill_type="solid", fgColor="F3F8F5")
        summary_font = Font(bold=True)

        for sheet_preview in workbook_preview.sheets:
            worksheet = workbook.create_sheet(title=sheet_preview.sheet_name[:31] or "Sheet")

            for column_index, header in enumerate(sheet_preview.headers, start=1):
                cell = worksheet.cell(row=1, column=column_index, value=header)
                cell.fill = header_fill
                cell.font = header_font
                worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(len(header) + 4, 14), 28)

            for row_index, row in enumerate(sheet_preview.rows, start=2):
                for column_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=column_index, value=value)

            if sheet_preview.summary_row:
                summary_row_index = len(sheet_preview.rows) + 2
                for column_index, value in enumerate(sheet_preview.summary_row, start=1):
                    cell = worksheet.cell(row=summary_row_index, column=column_index, value=value)
                    cell.fill = summary_fill
                    cell.font = summary_font

            worksheet.freeze_panes = "A2"

        note_sheet = workbook.create_sheet(title="说明")
        note_sheet["A1"] = "导出说明"
        note_sheet["A1"].font = Font(bold=True)
        current_row = 2
        for sheet_preview in workbook_preview.sheets:
            note_sheet.cell(row=current_row, column=1, value=f"{sheet_preview.sheet_name}")
            note_sheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            for note in sheet_preview.notes:
                note_sheet.cell(row=current_row, column=1, value=note)
                current_row += 1
            current_row += 1
        note_sheet.column_dimensions["A"].width = 120

        file_name = self._build_file_name(workbook_preview)
        file_path = self.export_dir / file_name
        workbook.save(file_path)
        logger.info(
            "Export Excel file created. file_path=%s parser_config_name=%s template_rule_name=%s batch_code=%s sheets=%s",
            file_path,
            workbook_preview.parser_config_name,
            workbook_preview.template_rule_name,
            workbook_preview.import_batch_code,
            len(workbook_preview.sheets),
        )
        return file_path

    def _build_file_name(self, workbook_preview: WorkbookPreview) -> str:
        base_name = f"{workbook_preview.parser_config_name}-{workbook_preview.template_rule_name}-{workbook_preview.import_batch_code}"
        safe_name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fa5_-]+", "_", base_name).strip("_") or "export"
        return f"{safe_name}.xlsx"
